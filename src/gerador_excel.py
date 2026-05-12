import os
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def ajustar_largura_colunas(ws):
    for coluna in ws.columns:
        maior_tamanho = 0
        letra_coluna = get_column_letter(coluna[0].column)

        for celula in coluna:
            if celula.value is not None:
                maior_tamanho = max(maior_tamanho, len(str(celula.value)))

        ws.column_dimensions[letra_coluna].width = min(maior_tamanho + 2, 45)


def formatar_cabecalho(ws):
    preenchimento = PatternFill(
        start_color="1F4E79",
        end_color="1F4E79",
        fill_type="solid"
    )

    fonte = Font(color="FFFFFF", bold=True)

    for celula in ws[1]:
        celula.fill = preenchimento
        celula.font = fonte
        celula.alignment = Alignment(horizontal="center", vertical="center")


def aplicar_bordas(ws):
    borda = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )

    for linha in ws.iter_rows():
        for celula in linha:
            celula.border = borda
            celula.alignment = Alignment(vertical="center")


def formatar_abas_excel(caminho_excel):
    wb = load_workbook(caminho_excel)

    for ws in wb.worksheets:
        formatar_cabecalho(ws)
        aplicar_bordas(ws)
        ajustar_largura_colunas(ws)
        ws.freeze_panes = "A2"

    wb.save(caminho_excel)


def destacar_risco(caminho_excel, config):
    wb = load_workbook(caminho_excel)

    if "Estudantes em Risco" not in wb.sheetnames:
        wb.save(caminho_excel)
        return

    ws = wb["Estudantes em Risco"]

    cor_moderado = config["excel"]["cor_risco_moderado"]
    cor_elevado = config["excel"]["cor_risco_elevado"]

    headers = [cell.value for cell in ws[1]]

    if "classificacao_risco" not in headers:
        wb.save(caminho_excel)
        return

    col_risco = headers.index("classificacao_risco") + 1

    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=col_risco).value

        if valor == "Risco moderado":
            fill = PatternFill(
                start_color=cor_moderado,
                end_color=cor_moderado,
                fill_type="solid"
            )
        elif valor == "Risco elevado":
            fill = PatternFill(
                start_color=cor_elevado,
                end_color=cor_elevado,
                fill_type="solid"
            )
        else:
            fill = None

        if fill:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    wb.save(caminho_excel)


def destacar_bonificacao(caminho_excel, config):
    wb = load_workbook(caminho_excel)

    if "Bonificacao" not in wb.sheetnames:
        wb.save(caminho_excel)
        return

    ws = wb["Bonificacao"]

    cor_bonus = config["excel"]["cor_bonus"]

    headers = [cell.value for cell in ws[1]]

    if "bonus_total" not in headers:
        wb.save(caminho_excel)
        return

    col_bonus = headers.index("bonus_total") + 1

    for row in range(2, ws.max_row + 1):
        valor = ws.cell(row=row, column=col_bonus).value

        try:
            valor_float = float(valor)
        except Exception:
            valor_float = 0

        if valor_float > 0:
            fill = PatternFill(
                start_color=cor_bonus,
                end_color=cor_bonus,
                fill_type="solid"
            )

            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = fill

    wb.save(caminho_excel)


def gerar_excel(resultados, pasta_execucao, config):
    pasta_excel = os.path.join(pasta_execucao, "excel")
    os.makedirs(pasta_excel, exist_ok=True)

    caminho_excel = os.path.join(
        pasta_excel,
        "enamed_performance_analytics_resultados.xlsx"
    )

    with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
        resultados["dados_processados"].to_excel(
            writer,
            sheet_name="Dados Processados",
            index=False
        )

        pd.DataFrame([resultados["estatisticas_gerais"]]).to_excel(
            writer,
            sheet_name="Estatisticas Gerais",
            index=False
        )

        resultados["ranking_geral"].to_excel(
            writer,
            sheet_name="Ranking Geral",
            index=False
        )

        resultados["ranking_por_turma"].to_excel(
            writer,
            sheet_name="Ranking por Turma",
            index=False
        )

        resultados["analise_por_turma"].to_excel(
            writer,
            sheet_name="Analise por Turma",
            index=False
        )

        resultados["analise_por_ciclo"].to_excel(
            writer,
            sheet_name="Analise por Ciclo",
            index=False
        )

        resultados["analise_por_areas"].to_excel(
            writer,
            sheet_name="Analise por Areas",
            index=False
        )

        resultados["analise_areas_por_turma"].to_excel(
            writer,
            sheet_name="Areas por Turma",
            index=False
        )

        resultados["estudantes_em_risco"].to_excel(
            writer,
            sheet_name="Estudantes em Risco",
            index=False
        )

        resultados["bonificacao"].to_excel(
            writer,
            sheet_name="Bonificacao",
            index=False
        )

    formatar_abas_excel(caminho_excel)
    destacar_risco(caminho_excel, config)
    destacar_bonificacao(caminho_excel, config)

    return caminho_excel